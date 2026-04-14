from __future__ import annotations

from datetime import date
from datetime import datetime
from datetime import time
from pathlib import Path
from threading import RLock
from typing import Any
from zipfile import ZipFile
import xml.etree.ElementTree as ElementTree

from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries


EXCEL_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
EXCEL_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
X14_NS = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"
XM_NS = "http://schemas.microsoft.com/office/excel/2006/main"

XML_NAMESPACES = {
    "main": EXCEL_MAIN_NS,
    "rel": EXCEL_REL_NS,
    "pkgrel": PACKAGE_REL_NS,
    "x14": X14_NS,
    "xm": XM_NS,
}

_validation_cache_lock = RLock()
_sheet_openpyxl_validation_cache: dict[str, dict[str, dict[str, Any]]] = {}
_sheet_xml_validation_cache: dict[str, dict[str, dict[str, Any]]] = {}


def normalize_value(value: Any) -> Any:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.isoformat()

    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, time):
        return value.isoformat()

    if isinstance(value, Path):
        return str(value)

    return value


def get_merged_parent_cell(sheet, row: int, column: int):
    cell = sheet.cell(row=row, column=column)

    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return sheet.cell(row=merged_range.min_row, column=merged_range.min_col)

    return cell


def get_display_cells(formula_sheet, value_sheet, row: int, column: int):
    formula_cell = get_merged_parent_cell(formula_sheet, row, column)
    value_cell = get_merged_parent_cell(value_sheet, row, column)
    return formula_cell, value_cell


def extract_fill_color(cell) -> str | None:
    fill = getattr(cell, "fill", None)

    if fill is None:
        return None

    if getattr(fill, "fill_type", None) is None:
        return None

    fg_color = getattr(fill, "fgColor", None)
    if fg_color is None:
        return None

    color_type = getattr(fg_color, "type", None)

    if color_type == "rgb":
        rgb = getattr(fg_color, "rgb", None)
        if isinstance(rgb, str):
            rgb_text = rgb.strip().upper()
            if rgb_text:
                if len(rgb_text) == 8:
                    return f"#{rgb_text[2:]}"
                if len(rgb_text) == 6:
                    return f"#{rgb_text}"

    if color_type == "theme":
        theme = getattr(fg_color, "theme", None)
        tint = getattr(fg_color, "tint", None)

        if theme is not None:
            if tint is not None and float(tint) != 0:
                return f"theme:{theme}:tint:{tint}"
            return f"theme:{theme}"

    if color_type == "indexed":
        indexed = getattr(fg_color, "indexed", None)
        if isinstance(indexed, int):
            return f"indexed:{indexed}"

    rgb = getattr(fg_color, "rgb", None)
    if isinstance(rgb, str):
        rgb_text = rgb.strip().upper()
        if rgb_text:
            if len(rgb_text) == 8:
                return f"#{rgb_text[2:]}"
            if len(rgb_text) == 6:
                return f"#{rgb_text}"

    theme = getattr(fg_color, "theme", None)
    if isinstance(theme, int):
        return f"theme:{theme}"

    indexed = getattr(fg_color, "indexed", None)
    if isinstance(indexed, int):
        return f"indexed:{indexed}"

    return None


def coordinate_in_sqref(coordinate: str, sqref: str) -> bool:
    try:
        from openpyxl.worksheet.cell_range import MultiCellRange

        multi_range = MultiCellRange(sqref)
        for cell_range in multi_range.ranges:
            if coordinate in cell_range:
                return True
    except Exception:
        pass

    return False


def _extract_xml_text(node) -> str | None:
    if node is None:
        return None

    text = getattr(node, "text", None)
    if text is None:
        return None

    stripped = text.strip()
    return stripped or None


def _normalize_formula_text(formula_text: str | None) -> str | None:
    if formula_text is None:
        return None

    stripped = formula_text.strip()
    if not stripped:
        return None

    return stripped


def _build_validation_info(
    validation_type: str | None,
    formula1: str | None,
    formula2: str | None,
    allow_blank: Any,
    show_drop_down: Any,
    show_input_message: Any,
    show_error_message: Any,
    sqref: str,
    source: str,
) -> dict[str, Any]:
    return {
        "type": validation_type,
        "formula1": str(formula1),
        "formula2": str(formula2),
        "allowBlank": allow_blank,
        "showDropDown": show_drop_down,
        "showInputMessage": show_input_message,
        "showErrorMessage": show_error_message,
        "sqref": sqref,
        "isListDropdown": validation_type == "list",
        "source": source,
    }


def _build_sheet_cache_key(sheet) -> str:
    workbook_path = getattr(getattr(sheet, "parent", None), "_hft_workbook_path", None)

    if workbook_path is None:
        return f"memory::{sheet.title}"

    workbook_path = Path(workbook_path)
    try:
        modified_time = workbook_path.stat().st_mtime
    except OSError:
        modified_time = 0

    return f"{workbook_path.resolve()}::{modified_time}::{sheet.title}"


def _get_sheet_relationship_target_map(workbook_path: Path) -> dict[str, str]:
    relationships_map: dict[str, str] = {}

    with ZipFile(workbook_path, "r") as archive:
        workbook_xml = archive.read("xl/workbook.xml")
        workbook_root = ElementTree.fromstring(workbook_xml)

        workbook_rels_xml = archive.read("xl/_rels/workbook.xml.rels")
        workbook_rels_root = ElementTree.fromstring(workbook_rels_xml)

        rel_id_to_target: dict[str, str] = {}
        for relationship in workbook_rels_root.findall("pkgrel:Relationship", XML_NAMESPACES):
            relationship_id = relationship.attrib.get("Id")
            target = relationship.attrib.get("Target")

            if relationship_id and target:
                rel_id_to_target[relationship_id] = target

        sheets_node = workbook_root.find("main:sheets", XML_NAMESPACES)
        if sheets_node is None:
            return relationships_map

        for sheet_node in sheets_node.findall("main:sheet", XML_NAMESPACES):
            sheet_name = sheet_node.attrib.get("name")
            relationship_id = sheet_node.attrib.get(f"{{{EXCEL_REL_NS}}}id")

            if not sheet_name or not relationship_id:
                continue

            target = rel_id_to_target.get(relationship_id)
            if not target:
                continue

            normalized_target = target.replace("\\", "/")
            if not normalized_target.startswith("worksheets/"):
                continue

            relationships_map[sheet_name] = f"xl/{normalized_target}"

    return relationships_map


def _get_sheet_xml_path(workbook_path: Path, sheet_name: str) -> str | None:
    try:
        mapping = _get_sheet_relationship_target_map(workbook_path)
        return mapping.get(sheet_name)
    except Exception:
        return None


def _read_standard_xml_validations(workbook_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    xml_path = _get_sheet_xml_path(workbook_path, sheet_name)
    if xml_path is None:
        return []

    validations: list[dict[str, Any]] = []

    with ZipFile(workbook_path, "r") as archive:
        if xml_path not in archive.namelist():
            return []

        sheet_xml = archive.read(xml_path)
        root = ElementTree.fromstring(sheet_xml)

        data_validations_node = root.find("main:dataValidations", XML_NAMESPACES)
        if data_validations_node is None:
            return []

        for validation_node in data_validations_node.findall("main:dataValidation", XML_NAMESPACES):
            sqref = (validation_node.attrib.get("sqref") or "").strip()
            if not sqref:
                continue

            validation_type = validation_node.attrib.get("type")
            formula1 = _normalize_formula_text(
                _extract_xml_text(validation_node.find("main:formula1", XML_NAMESPACES))
            )
            formula2 = _normalize_formula_text(
                _extract_xml_text(validation_node.find("main:formula2", XML_NAMESPACES))
            )

            validations.append(
                _build_validation_info(
                    validation_type=validation_type,
                    formula1=formula1,
                    formula2=formula2,
                    allow_blank=validation_node.attrib.get("allowBlank"),
                    show_drop_down=validation_node.attrib.get("showDropDown"),
                    show_input_message=validation_node.attrib.get("showInputMessage"),
                    show_error_message=validation_node.attrib.get("showErrorMessage"),
                    sqref=sqref,
                    source="xml-standard",
                )
            )

    return validations


def _read_x14_xml_validations(workbook_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    xml_path = _get_sheet_xml_path(workbook_path, sheet_name)
    if xml_path is None:
        return []

    validations: list[dict[str, Any]] = []

    with ZipFile(workbook_path, "r") as archive:
        if xml_path not in archive.namelist():
            return []

        sheet_xml = archive.read(xml_path)
        root = ElementTree.fromstring(sheet_xml)

        ext_list_node = root.find("main:extLst", XML_NAMESPACES)
        if ext_list_node is None:
            return []

        for ext_node in ext_list_node.findall("main:ext", XML_NAMESPACES):
            data_validations_node = ext_node.find("x14:dataValidations", XML_NAMESPACES)
            if data_validations_node is None:
                continue

            for validation_node in data_validations_node.findall("x14:dataValidation", XML_NAMESPACES):
                sqref_parts: list[str] = []

                for sqref_node in validation_node.findall("x14:sqref", XML_NAMESPACES):
                    sqref_text = _extract_xml_text(sqref_node)
                    if sqref_text:
                        sqref_parts.append(sqref_text)

                if not sqref_parts:
                    for xm_sqref_node in validation_node.findall("xm:sqref", XML_NAMESPACES):
                        sqref_text = _extract_xml_text(xm_sqref_node)
                        if sqref_text:
                            sqref_parts.append(sqref_text)

                sqref = " ".join(part.strip() for part in sqref_parts if part.strip())
                if not sqref:
                    continue

                formula1 = None
                formula2 = None

                formula1_node = validation_node.find("x14:formula1", XML_NAMESPACES)
                if formula1_node is not None:
                    formula1 = _extract_xml_text(formula1_node.find("xm:f", XML_NAMESPACES))
                    if formula1 is None:
                        formula1 = _extract_xml_text(formula1_node)

                formula2_node = validation_node.find("x14:formula2", XML_NAMESPACES)
                if formula2_node is not None:
                    formula2 = _extract_xml_text(formula2_node.find("xm:f", XML_NAMESPACES))
                    if formula2 is None:
                        formula2 = _extract_xml_text(formula2_node)

                validation_type = validation_node.attrib.get("type")

                validations.append(
                    _build_validation_info(
                        validation_type=validation_type,
                        formula1=_normalize_formula_text(formula1),
                        formula2=_normalize_formula_text(formula2),
                        allow_blank=validation_node.attrib.get("allowBlank"),
                        show_drop_down=validation_node.attrib.get("showDropDown"),
                        show_input_message=validation_node.attrib.get("showInputMessage"),
                        show_error_message=validation_node.attrib.get("showErrorMessage"),
                        sqref=sqref,
                        source="xml-x14",
                    )
                )

    return validations


def _build_openpyxl_validation_map(sheet) -> dict[str, dict[str, Any]]:
    validation_map: dict[str, dict[str, Any]] = {}

    data_validations = getattr(sheet, "data_validations", None)
    if not data_validations:
        return validation_map

    validations = getattr(data_validations, "dataValidation", None)
    if not validations:
        return validation_map

    for validation in validations:
        sqref = str(getattr(validation, "sqref", "") or "").strip()
        if not sqref:
            continue

        validation_info = _build_validation_info(
            validation_type=getattr(validation, "type", None),
            formula1=getattr(validation, "formula1", None),
            formula2=getattr(validation, "formula2", None),
            allow_blank=getattr(validation, "allowBlank", None),
            show_drop_down=getattr(validation, "showDropDown", None),
            show_input_message=getattr(validation, "showInputMessage", None),
            show_error_message=getattr(validation, "showErrorMessage", None),
            sqref=sqref,
            source="openpyxl",
        )

        for address in expand_sqref_to_addresses(sqref):
            validation_map[address] = validation_info

    return validation_map


def _build_xml_validation_map(sheet) -> dict[str, dict[str, Any]]:
    workbook_path = getattr(getattr(sheet, "parent", None), "_hft_workbook_path", None)
    if workbook_path is None:
        return {}

    workbook_path = Path(workbook_path)

    validations: list[dict[str, Any]] = []
    validations.extend(_read_standard_xml_validations(workbook_path, sheet.title))
    validations.extend(_read_x14_xml_validations(workbook_path, sheet.title))

    validation_map: dict[str, dict[str, Any]] = {}

    for validation in validations:
        sqref = validation.get("sqref")
        if not sqref:
            continue

        for address in expand_sqref_to_addresses(sqref):
            validation_map[address] = validation

    return validation_map


def _get_openpyxl_validation_map_cached(sheet) -> dict[str, dict[str, Any]]:
    cache_key = _build_sheet_cache_key(sheet)

    with _validation_cache_lock:
        cached = _sheet_openpyxl_validation_cache.get(cache_key)
        if cached is not None:
            return cached

        built = _build_openpyxl_validation_map(sheet)
        _sheet_openpyxl_validation_cache[cache_key] = built
        return built


def _get_xml_validation_map_cached(sheet) -> dict[str, dict[str, Any]]:
    cache_key = _build_sheet_cache_key(sheet)

    with _validation_cache_lock:
        cached = _sheet_xml_validation_cache.get(cache_key)
        if cached is not None:
            return cached

        built = _build_xml_validation_map(sheet)
        _sheet_xml_validation_cache[cache_key] = built
        return built


def get_data_validation_info(sheet, coordinate: str) -> dict[str, Any] | None:
    openpyxl_validation_map = _get_openpyxl_validation_map_cached(sheet)
    validation_info = openpyxl_validation_map.get(coordinate)
    if validation_info is not None:
        return validation_info

    xml_validation_map = _get_xml_validation_map_cached(sheet)
    return xml_validation_map.get(coordinate)


def get_all_list_dropdowns(sheet) -> list[dict[str, Any]]:
    unique_items: dict[str, dict[str, Any]] = {}

    for address, validation in _get_openpyxl_validation_map_cached(sheet).items():
        if validation.get("type") != "list":
            continue

        key = f"openpyxl::{validation.get('sqref')}::{validation.get('formula1')}::{address}"
        unique_items[key] = validation

    for address, validation in _get_xml_validation_map_cached(sheet).items():
        if validation.get("type") != "list":
            continue

        key = f"xml::{validation.get('sqref')}::{validation.get('formula1')}::{address}"
        unique_items[key] = validation

    return list(unique_items.values())


def expand_sqref_to_addresses(sqref: str) -> list[str]:
    addresses: list[str] = []

    for part in sqref.split():
        if ":" in part:
            min_column, min_row, max_column, max_row = range_boundaries(part)
            for row_index in range(min_row, max_row + 1):
                for column_index in range(min_column, max_column + 1):
                    addresses.append(f"{get_column_letter(column_index)}{row_index}")
        else:
            addresses.append(part)

    return addresses


def has_data_validation(sheet, coordinate: str) -> bool:
    return get_data_validation_info(sheet, coordinate) is not None


def has_list_dropdown(sheet, coordinate: str) -> bool:
    validation_info = get_data_validation_info(sheet, coordinate)
    if validation_info is None:
        return False
    return validation_info.get("type") == "list"


def classify_cell(formula_cell, value_cell, sheet, requested_coordinate: str) -> dict[str, bool | str | None | dict[str, Any]]:
    fill_color = extract_fill_color(formula_cell)
    has_formula = formula_cell.data_type == "f"

    validation_info = get_data_validation_info(sheet, requested_coordinate)
    has_validation = validation_info is not None
    has_dropdown = bool(validation_info and validation_info.get("type") == "list")

    formula_value = formula_cell.value
    cached_value = value_cell.value

    is_input_candidate = (
        not has_formula
        and (formula_value is not None or cached_value is not None or has_dropdown)
        and (has_dropdown or fill_color is not None)
    )

    is_result_candidate = has_formula

    return {
        "hasFormula": has_formula,
        "hasValidation": has_validation,
        "hasDropdown": has_dropdown,
        "validationInfo": validation_info,
        "isInputCandidate": is_input_candidate,
        "isResultCandidate": is_result_candidate,
        "fillColor": fill_color,
    }


def build_cell_payload(
    formula_cell,
    value_cell,
    sheet,
    requested_coordinate: str,
    include_empty: bool = False,
) -> dict[str, Any] | None:
    formula_value = formula_cell.value
    cached_value = value_cell.value

    classification = classify_cell(
        formula_cell=formula_cell,
        value_cell=value_cell,
        sheet=sheet,
        requested_coordinate=requested_coordinate,
    )

    if (
        not include_empty
        and formula_value is None
        and cached_value is None
        and not classification["hasDropdown"]
    ):
        return None

    return {
        "address": formula_cell.coordinate,
        "row": formula_cell.row,
        "column": formula_cell.column,
        "value": normalize_value(formula_value),
        "formula": formula_value if formula_cell.data_type == "f" else None,
        "cachedValue": normalize_value(cached_value),
        "dataType": formula_cell.data_type,
        "numberFormat": formula_cell.number_format,
        "fillColor": classification["fillColor"],
        "hasFormula": classification["hasFormula"],
        "hasValidation": classification["hasValidation"],
        "hasDropdown": classification["hasDropdown"],
        "validationInfo": classification["validationInfo"],
        "isInputCandidate": classification["isInputCandidate"],
        "isResultCandidate": classification["isResultCandidate"],
    }