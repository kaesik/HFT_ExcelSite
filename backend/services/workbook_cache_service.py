from __future__ import annotations

from pathlib import Path
from threading import RLock
from typing import Any

from openpyxl import load_workbook


class WorkbookCacheService:
    def __init__(self) -> None:
        self._lock = RLock()
        self._cache: dict[str, dict[str, Any]] = {}

    def _build_cache_key(self, workbook_path: Path) -> str:
        return str(workbook_path.resolve()).lower()

    def _get_mtime(self, workbook_path: Path) -> float:
        return workbook_path.stat().st_mtime

    def _load_formula_workbook(self, workbook_path: Path):
        workbook = load_workbook(
            filename=workbook_path,
            data_only=False,
            read_only=False,
            keep_vba=True,
        )
        setattr(workbook, "_hft_workbook_path", workbook_path)
        return workbook

    def _load_value_workbook(self, workbook_path: Path):
        workbook = load_workbook(
            filename=workbook_path,
            data_only=True,
            read_only=False,
            keep_vba=True,
        )
        setattr(workbook, "_hft_workbook_path", workbook_path)
        return workbook

    def _create_entry(self, workbook_path: Path) -> dict[str, Any]:
        return {
            "mtime": self._get_mtime(workbook_path),
            "formula_workbook": self._load_formula_workbook(workbook_path),
            "value_workbook": self._load_value_workbook(workbook_path),
        }

    def get_formula_workbook(self, workbook_path: Path):
        key = self._build_cache_key(workbook_path)
        current_mtime = self._get_mtime(workbook_path)

        with self._lock:
            entry = self._cache.get(key)

            if entry is None or entry["mtime"] != current_mtime:
                entry = self._create_entry(workbook_path)
                self._cache[key] = entry

            return entry["formula_workbook"]

    def get_value_workbook(self, workbook_path: Path):
        key = self._build_cache_key(workbook_path)
        current_mtime = self._get_mtime(workbook_path)

        with self._lock:
            entry = self._cache.get(key)

            if entry is None or entry["mtime"] != current_mtime:
                entry = self._create_entry(workbook_path)
                self._cache[key] = entry

            return entry["value_workbook"]

    def get_two_workbooks(self, workbook_path: Path):
        key = self._build_cache_key(workbook_path)
        current_mtime = self._get_mtime(workbook_path)

        with self._lock:
            entry = self._cache.get(key)

            if entry is None or entry["mtime"] != current_mtime:
                entry = self._create_entry(workbook_path)
                self._cache[key] = entry

            return entry["formula_workbook"], entry["value_workbook"]

    def invalidate(self, workbook_path: Path) -> None:
        key = self._build_cache_key(workbook_path)

        with self._lock:
            self._cache.pop(key, None)

    def clear(self) -> None:
        with self._lock:
            self._cache.clear()


workbook_cache_service = WorkbookCacheService()