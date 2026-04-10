from pathlib import Path

from fastapi import HTTPException
from openpyxl import load_workbook

from app.config import ALLOWED_EXTENSIONS
from app.config import WORKBOOKS_DIR


def ensure_workbooks_dir_exists() -> None:
    if not WORKBOOKS_DIR.exists():
        raise FileNotFoundError(f"Workbooks directory not found: {WORKBOOKS_DIR}")


def list_workbook_files() -> list[Path]:
    ensure_workbooks_dir_exists()

    files = [
        path
        for path in WORKBOOKS_DIR.iterdir()
        if path.is_file()
        and path.suffix.lower() in ALLOWED_EXTENSIONS
        and not path.name.startswith("~$")
    ]

    files.sort(key=lambda item: item.name.lower())
    return files


def get_workbook_path(file_name: str) -> Path:
    ensure_workbooks_dir_exists()

    workbook_path = WORKBOOKS_DIR / file_name

    if workbook_path.name.startswith("~$"):
        raise HTTPException(status_code=400, detail="Temporary Excel lock files are not supported")

    if workbook_path.suffix.lower() not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Unsupported workbook extension")

    if not workbook_path.exists() or not workbook_path.is_file():
        raise HTTPException(status_code=404, detail=f"Workbook '{file_name}' not found")

    return workbook_path


def load_formula_workbook(workbook_path: Path):
    return load_workbook(
        filename=workbook_path,
        data_only=False,
        read_only=False,
        keep_vba=True,
    )


def load_value_workbook(workbook_path: Path):
    return load_workbook(
        filename=workbook_path,
        data_only=True,
        read_only=False,
        keep_vba=True,
    )


def get_two_workbooks(workbook_path: Path):
    workbook_with_formulas = load_formula_workbook(workbook_path)
    workbook_with_values = load_value_workbook(workbook_path)
    return workbook_with_formulas, workbook_with_values


def validate_sheet_name(workbook, sheet_name: str) -> None:
    if sheet_name not in workbook.sheetnames:
        raise HTTPException(
            status_code=404,
            detail=f"Sheet '{sheet_name}' not found",
        )